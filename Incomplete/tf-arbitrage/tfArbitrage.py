import os
import re
import time
import requests
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
import steampy
from steampy.guard import generate_one_time_code
from steampy.client import SteamClient
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import selenium
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


#Global Variables
#User Agent Variables
ua = UserAgent()
headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_4) AppleWebKit/537.13 (KHTML, like Gecko) Chrome/24.0.1290.1 Safari/537.13'}
sheetDirectory = ".\tfArbitrage.xlsx"

#SteamPy Variables
#Read info.txt
infoDirectory = ".\info.txt"
if(os.path.exists(infoDirectory) != True):
    infoCreate()

#TODO Check for llines in text doc
#Load variables from info file
infoFile = open(".\info.txt", "r")
infoArray = infoFile.readlines()
# Set API
api_key = infoArray[1][:-1]
# Steam steamname
steamname = str(infoArray[3][:-1])
# Steam password
password = str(infoArray[5][:-1])
# steam's Shared Secret
secret = infoArray[7]
delay = 5



#Function definitions
#Creates info.txt file
def infoCreate():
    infoFile = open(infoDirectory, "w+")
    infoFile.write('''#Put steam API key on line below

#Put Username on line below

#Put Password on line below

#Put shared secret on line below
 ''')
    infoFile.close()
    print('--File created at ' + infoDirectory)
    incompleteInfo()

def incompleteInfo():
    print('Txt file missing information. Please fill it out.')
    sys.quit()

def getAuthCode():
    steamPyAuthCode = generate_one_time_code(secret)
    return steamPyAuthCode

    #Logs into steam, given correct authcode
def seleniumLogin(authCode):

    #Browser setup
    browser = webdriver.Chrome()
    browser.maximize_window()
    browser.get("https://scrap.tf/buy/hats")
    wait = WebDriverWait(browser, delay)
    wait.until(EC.element_to_be_clickable((By.ID, "steamAccountName"))).send_keys(steamname)
    wait.until(EC.element_to_be_clickable((By.ID, "steamPassword"))).send_keys(password)
    wait.until(EC.element_to_be_clickable((By.ID, "steamPassword"))).send_keys(Keys.RETURN)
    wait.until(EC.element_to_be_clickable((By.ID, "twofactorcode_entry"))).send_keys(authCode)
    wait.until(EC.element_to_be_clickable((By.ID, "twofactorcode_entry"))).send_keys(Keys.RETURN)

    #Variable declaration
    elements = []
    file = 'tfArbitrage.xlsx'
    wb = load_workbook(filename=file)
    ws = wb.active

    #Scraping the site for items it has available
    time.sleep(delay)
    #Use soup on finished page
    browser.get("https://scrap.tf/buy/hats")
    soup = BeautifulSoup(browser.page_source, "html.parser")
    elementContainers = soup.find_all('div', class_="items-container")
    for container in elementContainers:#Site has all items in divs in item container classes, this selects all of them
        for element in container.find_all('div'):#Tracks down all divs in the results
            dataID = element.get('data-id')#Only processes divs with a data-id attribute
            if (dataID != None and element.get('data-content') != "&lt;b&gt;This item is overstocked and cannot be sold.&lt;/b&gt;"):
                elemData = [] #Refresh elemData variable, to store information on each specific item
                #Check if item is available
                itemQuantity = element.get('data-bot23-count')

                elemData.append(dataID)#Item ID Number
                elemData.append(element.get('data-title'))#Item Name
                elemData.append(element['class'][2][7:])#Item Quality Number
                elemData.append(element.get('data-bot23-count'))#Num of item available

                #Block of code to use regex to sort through information in HTML block pulled
                itemCost = element.get('data-content')
                regex = re.compile(r"\s[\d\s\w.,]*refined")
                itemCost = regex.findall(itemCost)
                elemData.append(itemCost)#Item content, incl. Cost
                #if elemData[4] != '':
                #    elements.append(elemData) #Should select each item



    #declare info about the sheet here
    for i in range(0, len(elements)):
        for j in range(0, len(elements[i])):
          ws.cell(row=i+2, column=j+1).value = str(elements[i][j])#Add 1 to i, so it has space to work with

    wb.save(file)

#Creates spreadsheet for data storage
def sheetCreate():
    print("--Creating tfArbitrage.xlsx in current directory")
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "Item ID"
    sheet["B1"] = "Name"
    sheet["C1"] = "Item Quality"
    sheet["D1"] = "Quantity"
    sheet["E1"] = "Price"

    workbook.save(filename="tfArbitrage.xlsx")

    progStart()

'''
#CURRENTLY UNIMPLEMENTED
#Will create listings for each item considered "Viable" on the spreadsheet
def listingPosts():
    print("-- Posting listings for inventory.")
    tradeBot()

#Will respond to appropriate messages on steam
    #Should be running consistently, as the rest of the program is running. Or run itermittently as it checks for messages.
def tradeBot():
    print("-- Managing incoming trade offers and messages.")
    listingPosts()
'''

def progStart():
    authCode = getAuthCode()
    seleniumLogin(authCode)

def main():
    if(os.path.exists(sheetDirectory)):
        progStart()
    else:
        sheetCreate()


main()
